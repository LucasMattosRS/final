import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

# ── Estilos ───────────────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
_OK_FILL      = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_OK_FONT      = Font(color="276221", bold=True)
_PEND_FILL    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_PEND_FONT    = Font(color="9C6500", bold=True)
_CRIT_FILL = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

_CRIT_FONT = Font(
    color="9C0006",
    bold=True
)
_THIN_BORDER  = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Colunas na ordem desejada (só as que existirem no DataFrame serão incluídas)
_COLUMN_ORDER = [
    "obra",
    "folha",
    "tipo",
    "codigo",
    "logradouro",
    "ancoragem",
    "lado_forte",
    "metragem",
    "material",
    "altura_poste",
    "informacoes",
    "confianca",
    "status",
    "corrigido",
]
_COLUMN_LABELS = {
    "obra":         "Obra",
    "folha":        "Folha",
    "tipo":         "Tipo",
    "codigo":       "Código",
    "logradouro":   "Rua / Avenida",
    "ancoragem":    "Ancoragem",
    "lado_forte":   "Lado Forte",
    "metragem":     "Metragem",
    "material":     "Material",
    "altura_poste": "Altura Poste",
    "informacoes":  "Informações",
    "confianca":    "Confiança",
    "status":       "Status",
    "corrigido":    "Corrigido IA",
}


def export_excel(data: list[dict], output_file: str) -> None:
    if not data:
        logger.warning("Sem dados para exportar em %s", output_file)
        return

    df = pd.DataFrame(data)
    colunas = [c for c in _COLUMN_ORDER if c in df.columns]
    df = df[colunas]

    # Renomear colunas para português
    df = df.rename(columns=_COLUMN_LABELS)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Resultado", index=False)
        ws = writer.sheets["Resultado"]

        # ── Cabeçalho ─────────────────────────────────────────────────────
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER

        # ── Linhas de dados ───────────────────────────────────────────────
        conf_col_idx = None
        status_col_idx = None

        if "Confiança" in df.columns:
            conf_col_idx = df.columns.get_loc("Confiança") + 1

        if "Status" in df.columns:
            status_col_idx = df.columns.get_loc("Status") + 1

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = _THIN_BORDER

                # Coluna Confiança
                if conf_col_idx and cell.column == conf_col_idx:
                    if cell.value >= 80:
                        cell.fill = _OK_FILL
                        cell.font = _OK_FONT
                    elif cell.value >= 60:
                        cell.fill = _PEND_FILL
                        cell.font = _PEND_FONT
                    else:
                        cell.fill = _CRIT_FILL
                        cell.font = _CRIT_FONT
                    cell.alignment = _CENTER

                # Coluna Status
                elif status_col_idx and cell.column == status_col_idx:
                    if cell.value == "APROVADO":
                        cell.fill = _OK_FILL
                        cell.font = _OK_FONT
                    elif cell.value == "REVISAR":
                        cell.fill = _PEND_FILL
                        cell.font = _PEND_FONT
                    elif cell.value == "CRITICO":
                        cell.fill = _CRIT_FILL
                        cell.font = _CRIT_FONT
                    cell.alignment = _CENTER

                else:
                    cell.alignment = _LEFT

        # ── Largura automática das colunas ────────────────────────────────
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=8,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

        # ── Congelar linha do cabeçalho ───────────────────────────────────
        ws.freeze_panes = "A2"

    logger.info("Excel gerado: %s (%d linhas)", output_file, len(df))
