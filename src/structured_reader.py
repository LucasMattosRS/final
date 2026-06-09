"""
src/structured_reader.py
────────────────────────
ITEM 4 — Teto de precisão.

Se o software de projeto (CAD/AutoCAD/sistema da distribuidora) conseguir
exportar os dados em planilha/CSV, este leitor ingere esse arquivo e produz
diretamente as linhas no MESMO formato que o pipeline de OCR gera — pulando
toda a etapa de imagem/seta/heurística. Precisão ~100%, pois não há leitura
visual envolvida.

Formato esperado (colunas mínimas; nomes flexíveis — veja MAPA_COLUNAS):
    obra, folha, tipo, codigo, logradouro, ancoragem, lado_forte,
    metragem, material, altura_poste

Aceita .csv e .xlsx. Colunas ausentes viram vazio. Colunas extras são ignoradas.

Uso no código:
    from src.structured_reader import importar_estruturado
    linhas = importar_estruturado("export_obra.xlsx")
    export_excel(linhas, "saida.xlsx")
"""
import os
import csv

# aceita variações de nome de coluna (tudo comparado em minúsculo, sem espaços)
MAPA_COLUNAS = {
    "obra": ["obra", "ordem", "os", "numero_obra", "n_obra"],
    "folha": ["folha", "sheet", "pagina"],
    "tipo": ["tipo", "elemento"],
    "codigo": ["codigo", "código", "id", "tag", "poste", "vao", "vão"],
    "logradouro": ["logradouro", "rua", "rua_avenida", "endereco", "endereço", "via"],
    "ancoragem": ["ancoragem", "ancora", "âncora", "estai"],
    "lado_forte": ["lado_forte", "lado forte", "ladoforte", "lado"],
    "metragem": ["metragem", "vao_m", "comprimento", "distancia", "distância"],
    "material": ["material", "poste_material", "tipo_poste"],
    "altura_poste": ["altura_poste", "altura", "altura_m"],
}

SAIDA_COLS = list(MAPA_COLUNAS.keys())


def _normaliza_cabecalho(nome: str) -> str:
    return str(nome or "").strip().lower().replace("  ", " ")


def _mapear(cabecalhos: list[str]) -> dict:
    """Liga cada coluna de saída ao índice/nome real no arquivo importado."""
    achados = {}
    norm = {_normaliza_cabecalho(c): c for c in cabecalhos}
    for destino, aliases in MAPA_COLUNAS.items():
        for a in aliases:
            if a in norm:
                achados[destino] = norm[a]
                break
    return achados


def _linhas_de_csv(caminho: str) -> list[dict]:
    with open(caminho, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _linhas_de_xlsx(caminho: str) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    cabec = [str(c) if c is not None else "" for c in next(rows)]
    out = []
    for r in rows:
        out.append({cabec[i]: (r[i] if i < len(r) else None) for i in range(len(cabec))})
    return out


def importar_estruturado(caminho: str) -> list[dict]:
    ext = os.path.splitext(caminho)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        brutas = _linhas_de_xlsx(caminho)
    elif ext in (".csv", ".tsv"):
        brutas = _linhas_de_csv(caminho)
    else:
        raise ValueError(f"Formato não suportado: {ext} (use .csv ou .xlsx)")

    if not brutas:
        return []

    mapa = _mapear(list(brutas[0].keys()))
    if "codigo" not in mapa:
        raise ValueError(
            "Não encontrei a coluna de código (poste/vão). "
            f"Cabeçalhos vistos: {list(brutas[0].keys())}"
        )

    saida = []
    for r in brutas:
        linha = {}
        for destino in SAIDA_COLS:
            origem = mapa.get(destino)
            val = r.get(origem, "") if origem else ""
            linha[destino] = "" if val is None else str(val).strip()
        # campos derivados de auditoria — vindo de fonte confiável
        linha["informacoes"] = ""
        linha["confianca"] = 100
        linha["status"] = "APROVADO"
        linha["corrigido"] = "NAO"
        if not linha["tipo"]:
            cod = linha["codigo"].upper()
            linha["tipo"] = "VAO" if cod.startswith("V") else "POSTE"
        saida.append(linha)
    return saida
